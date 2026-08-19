from pathlib import Path
from datetime import datetime
from collections import Counter
from urllib.parse import urlparse
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from 工具 import clean_email, normalize_phone

COLUMNS={"company_name":"公司名稱","category":"分類","email":"Email","phone":"電話","mobile":"手機","fax":"Fax",
         "address":"地址","city":"城市","province":"Province","website":"Website","facebook":"Facebook",
         "contact_person":"Contact Person","source":"來源","source_url":"來源URL","email_source_url":"Email來源URL",
         "completeness":"資料完整度"}


class Exporter:
    def __init__(self, database): self.database=database
    def _data(self): return self.database.all_companies()
    def export_excel(self, folder):
        folder=Path(folder); folder.mkdir(parents=True,exist_ok=True); stamp=datetime.now().strftime("%Y%m%d_%H%M%S")
        path=folder/f"Philippines_Construction_Companies_{stamp}.xlsx"; rows=self._data()
        frame=pd.DataFrame(rows).rename(columns=COLUMNS)
        wanted=list(COLUMNS.values()); frame=frame.reindex(columns=wanted)
        emails=sorted({e for row in rows for e in (clean_email(x) for x in (row.get("email") or "").split(";")) if e})
        phone_rows=[]
        for row in rows:
            for kind in ("phone","mobile"):
                for p in (row.get(kind) or "").split(";"):
                    if p.strip():phone_rows.append({"公司名稱":row["company_name"],"類型":"電話" if kind=="phone" else "手機","電話":p.strip()})
        with pd.ExcelWriter(path,engine="openpyxl") as writer:
            frame.to_excel(writer,sheet_name="全部公司",index=False)
            frame[frame["Email"].fillna("")!=""].to_excel(writer,sheet_name="有Email",index=False)
            frame[frame["Email"].fillna("")==""].to_excel(writer,sheet_name="無Email",index=False)
            pd.DataFrame({"Email":emails}).to_excel(writer,sheet_name="EMAIL",index=False)
            pd.DataFrame(phone_rows).to_excel(writer,sheet_name="電話",index=False)
            for sheet,col in (("來源統計","來源"),("分類統計","分類"),("城市統計","城市")):
                frame[col].fillna("").value_counts().rename_axis(col).reset_index(name="數量").to_excel(writer,sheet_name=sheet,index=False)
            domains=Counter(urlparse("//"+e.split("@")[-1]).netloc or e.split("@")[-1] for e in emails)
            pd.DataFrame(domains.items(),columns=["Email網域","數量"]).sort_values("數量",ascending=False).to_excel(writer,sheet_name="Email網域統計",index=False)
            for ws in writer.book.worksheets:
                ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
                for cell in ws[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="2563EB")
                for column in ws.columns:
                    width=min(60,max(10,max(len(str(c.value or "")) for c in column)+2)); ws.column_dimensions[get_column_letter(column[0].column)].width=width
                headers={c.value:c.column for c in ws[1]}
                for h in ("電話","手機","Fax"):
                    if h in headers:
                        for cell in list(ws.columns)[headers[h]-1][1:]:cell.number_format="@"
                for h in ("Email","Website","來源URL","Email來源URL"):
                    if h in headers:
                        for cell in list(ws.columns)[headers[h]-1][1:]:
                            value=str(cell.value or "")
                            if value and ";" not in value: cell.hyperlink=("mailto:"+value if h=="Email" else value); cell.style="Hyperlink"
        return path

    def export_txt(self, folder, kinds=None):
        folder=Path(folder);folder.mkdir(parents=True,exist_ok=True);stamp=datetime.now().strftime("%Y%m%d_%H%M%S"); rows=self._data(); out=[]
        kinds=set(kinds or ("full","email","phone","no_email","email_company"))
        emails=sorted({e for r in rows for e in (clean_email(x) for x in (r.get("email") or "").split(";")) if e})
        mapping={e:r["company_name"] for r in rows for e in (clean_email(x) for x in (r.get("email") or "").split(";")) if e}
        specs={
          "full":("Philippines_Construction_Companies_FULL_",[" | ".join(str(r.get(k,"") or "") for k in ("company_name","category","email","phone","mobile","address","city","province","website","source","source_url")) for r in rows]),
          "email":("Philippines_Construction_Emails_",emails),
          "phone":("Philippines_Construction_Phones_",sorted({n for r in rows for k in ("phone","mobile") for p in (r.get(k) or "").split(";") if (n:=normalize_phone(p))})),
          "no_email":("Philippines_Construction_NO_EMAIL_",[" | ".join(str(r.get(k,"") or "") for k in ("company_name","phone","address","city","website","source")) for r in rows if not r.get("email")]),
          "email_company":("Philippines_Construction_Email_Company_",[f"{e} | {mapping[e]}" for e in emails])}
        for kind in kinds:
            prefix,lines=specs[kind]; path=folder/f"{prefix}{stamp}.txt";path.write_text("\n".join(lines),encoding="utf-8-sig");out.append(path)
        return out

    def export_all(self, folder): return [self.export_excel(folder),*self.export_txt(folder)]

